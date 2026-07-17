"""
generar_acum_ant.py
-------------------
Arma el Excel comparativo del ANIO ANTERIOR para CUALQUIER rango de fechas,
leyendo el master fijo data/Ventas_Master_2025.xlsx.

Es la version generalizada de generar_acum2025.py:
  - El diario pide "del 1ro del mes hasta el dia X"  -> rango.
  - El semanal pide "del viernes al jueves"          -> rango.
Mismo calculo, distintos limites. Por eso una sola funcion.

CRITERIO DE ESPEJO: fechas calendario (decision de Juan).
  Rango 2026: 2026-07-17 .. 2026-07-23
  Rango 2025: 2025-07-17 .. 2025-07-23   <- mismo dia y mes, anio - 1
  OJO: NO alinea por dia de semana. El 17/07/2026 es viernes y el 17/07/2025 fue
  jueves, asi que la composicion de dias fuertes (sab/dom) puede no coincidir.
  Esta documentado a proposito: es una decision de negocio, no un bug.

Uso por linea de comando:
  python generar_acum_ant.py --desde 2026-07-17 --hasta 2026-07-23 \
                             --salida Acumulado_semanal_ant.xlsx

Uso como modulo (lo que hace report_semanal.py):
  from generar_acum_ant import acumular_rango, escribir_excel

Falla RUIDOSA: si al master le falta algun dia del rango, corta con error y te
dice cuales. Prefiero el job en rojo antes que un comparativo incompleto que le
baje la variacion a los socios sin que nadie se entere.
"""
import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
MASTER = BASE / "data" / "Ventas_Master_2025.xlsx"

# El master usa "00X - Nombre"; report.py espera los nombres estilo TouchBistro.
# Mapeo por numero de local para no depender de como este escrito el resto.
CODE_TO_TBKEY = {
    "001": "#001 Florida Mall Orlando FL",
    "002": "#002 American Dream Mall NJ",
    "003": "#003 Sawgrass Mills Mall FL",
    "004": "#004 Weston Town Center FL",
    "005": "#005 Vineland Orlando FL",
    "006": "#006 Aventura, FL",
}
ORDEN = ["001", "002", "003", "004", "005", "006"]


def _to_float(v):
    if v is None:
        return 0.0
    return float(str(v).replace(",", "").replace("$", "").strip() or 0)


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def espejo(d):
    """Misma fecha calendario del anio anterior.

    El 29/02 no existe en anios no bisiestos: en ese caso caigo al 28/02 en vez
    de reventar. Con el master actual (jun-dic 2025) no puede pasar, pero si
    manianа se carga enero-febrero, esto ya esta contemplado.
    """
    try:
        return d.replace(year=d.year - 1)
    except ValueError:
        return date(d.year - 1, 2, 28)


def _leer_master():
    wb = openpyxl.load_workbook(MASTER, data_only=True)
    ws = wb["Por Dia y Local"]
    header = [str(c) for c in next(ws.iter_rows(values_only=True))]
    col_local = header.index("Local")
    col_fecha = header.index("Fecha")
    col_sales = header.index("Sales")

    datos = {}  # {fecha: {cod: sales}}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[col_local] is None:
            continue
        cod = str(r[col_local]).strip()[:3]
        if cod not in CODE_TO_TBKEY:
            continue
        f = _as_date(r[col_fecha])
        datos.setdefault(f, {})[cod] = _to_float(r[col_sales])
    return datos


def acumular_rango(ini_ant, fin_ant):
    """Suma la columna Sales (Net Sales) por local entre ini_ant y fin_ant
    INCLUSIVE. Las fechas que se le pasan ya son del anio anterior.

    Devuelve (acum_por_codigo, dias_faltantes).
    """
    datos = _leer_master()

    acum = {cod: 0.0 for cod in ORDEN}
    faltantes = []
    d = ini_ant
    while d <= fin_ant:
        if d not in datos:
            faltantes.append(d)
        else:
            for cod, v in datos[d].items():
                acum[cod] += v
        d += timedelta(days=1)
    return acum, faltantes


def escribir_excel(ini_ant, fin_ant, acum, salida):
    """Escribe con el MISMO formato que TouchBistro para que parse_excel() de
    report.py lo lea sin cambiarle una coma:
      fila 0 = titulo con el rango, fila 1 = header, filas 2+ = venue y Net Sales en col C.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Summary"
    titulo = f"multi-venue - Sales Summary - {ini_ant.isoformat()}/{fin_ant.isoformat()}"
    ws.append([titulo, None, None, None, None, None])
    ws.append(["Venue Name", "Gross Sales", "Net Sales", "Discounts", "Voids", "Bill Count"])
    total = 0.0
    for cod in ORDEN:
        v = round(acum[cod], 2)
        total += v
        ws.append([CODE_TO_TBKEY[cod], v, v, 0, 0, 0])  # col C = Net Sales
    ws.append([f"REPORT SUMMARY ({len(ORDEN)} entries)", total, total, 0, 0, 0])
    wb.save(salida)
    return total


def main():
    ap = argparse.ArgumentParser(description="Comparativo del anio anterior por rango.")
    ap.add_argument("--desde", required=True, help="Fecha inicio del rango ACTUAL (YYYY-MM-DD)")
    ap.add_argument("--hasta", required=True, help="Fecha fin del rango ACTUAL (YYYY-MM-DD)")
    ap.add_argument("--salida", default="Acumulado_interanual.xlsx")
    a = ap.parse_args()

    desde = datetime.strptime(a.desde, "%Y-%m-%d").date()
    hasta = datetime.strptime(a.hasta, "%Y-%m-%d").date()
    if desde > hasta:
        print("[ERROR] --desde no puede ser posterior a --hasta.")
        return 1

    ini_ant, fin_ant = espejo(desde), espejo(hasta)
    acum, faltantes = acumular_rango(ini_ant, fin_ant)

    if faltantes:
        print(f"[ERROR] Al master 2025 le faltan {len(faltantes)} dia(s) del rango "
              f"{ini_ant}..{fin_ant}:")
        for d in faltantes:
            print(f"  - {d.isoformat()} ({d.strftime('%A')})")
        print("Cargalos en la hoja 'Por Dia y Local' antes de correr.")
        return 1

    total = escribir_excel(ini_ant, fin_ant, acum, BASE / a.salida)
    print(f"[OK] Comparativo {ini_ant}..{fin_ant} -> {a.salida} | Total: ${total:,.2f}")
    for cod in ORDEN:
        print(f"     {CODE_TO_TBKEY[cod]:<32} ${acum[cod]:>12,.2f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
