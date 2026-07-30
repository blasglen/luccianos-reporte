"""
mensual.py - Master mensual y metricas de largo plazo
-----------------------------------------------------
Fuente unica para:
  - Acumulado de los ULTIMOS 12 MESES moviles (lo usa el SEMANAL, columna por sucursal).
  - Acumulado del AÑO / YTD (lo usa el CIERRE, interanual).

Guarda un registro por mes y sucursal en data/mensual_master.json:
    { "2026-06": { "Florida Mall": {"venta": 73757.46, "tickets": 4998}, ... }, ... }

Como el resto del repo, PREFIERE FALLAR EN ROJO antes que devolver un numero
incompleto: si a una ventana le falta un mes, corta y dice cual.

Quien lo alimenta:
  - Backfill inicial: `python mensual.py backfill` (una sola vez).
  - En regimen: el CIERRE llama a registrar_mes() con el mes que acaba de cerrar,
    asi el master crece solo, un mes por vez, sin intervencion.

Control de dos caminos (mismo espiritu que la conciliacion del semanal): si un mes
esta tanto aca como en el historial diario, tienen que dar igual al centavo
(conciliar_con_historial), sino se aborta.
"""
import json
import sys
from datetime import date
from pathlib import Path

from report import BRANCH_ORDER

BASE = Path(__file__).parent
MASTER = BASE / "data" / "mensual_master.json"


# --- Lectura / escritura --------------------------------------------------------
def cargar():
    if not MASTER.exists():
        raise SystemExit(
            f"[ERROR] No existe {MASTER}. Corre el backfill una vez: "
            f"`python mensual.py backfill`."
        )
    return json.loads(MASTER.read_text(encoding="utf-8"))


def guardar(data):
    MASTER.parent.mkdir(parents=True, exist_ok=True)
    MASTER.write_text(
        json.dumps({k: data[k] for k in sorted(data)}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _clave(anio, mes):
    return f"{anio:04d}-{mes:02d}"


def meses_disponibles(data=None):
    data = data if data is not None else cargar()
    return sorted(tuple(map(int, k.split("-"))) for k in data)


def ultimo_mes(data=None):
    """El mes mas reciente cargado (== ultimo mes cerrado)."""
    return meses_disponibles(data)[-1]


# --- Registrar un mes (lo llama el cierre) --------------------------------------
def registrar_mes(anio, mes, por_sucursal, data=None):
    """por_sucursal = {suc: {'venta': x, 'tickets': y}}. Idempotente: la clave
    es el mes, asi que volver a registrar el mismo mes lo pisa con el mismo valor."""
    data = data if data is not None else cargar()
    faltan = [b for b in BRANCH_ORDER if b not in por_sucursal]
    if faltan:
        raise SystemExit(f"[ERROR] registrar_mes {_clave(anio,mes)}: faltan {faltan}")
    data[_clave(anio, mes)] = {
        b: {"venta": round(float(por_sucursal[b]["venta"]), 2),
            "tickets": int(por_sucursal[b]["tickets"])}
        for b in BRANCH_ORDER
    }
    guardar(data)
    return data


# --- Ventanas -------------------------------------------------------------------
def ventana_12m(anio, mes):
    """Los 12 (anio, mes) que terminan en (anio, mes), del mas viejo al mas nuevo."""
    out = []
    y, m = anio, mes
    for _ in range(12):
        out.append((y, m))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return list(reversed(out))


def _sumar(data, claves, contexto):
    venta = {b: 0.0 for b in BRANCH_ORDER}
    tickets = {b: 0 for b in BRANCH_ORDER}
    faltan = [k for k in claves if _clave(*k) not in data]
    if faltan:
        det = ", ".join(_clave(*k) for k in faltan)
        raise SystemExit(
            f"[ERROR] {contexto}: al master mensual le faltan {len(faltan)} mes(es): {det}.\n"
            f"        Backfill incompleto o el cierre de esos meses no corrio. "
            f"NO se manda un acumulado con huecos."
        )
    for k in claves:
        reg = data[_clave(*k)]
        for b in BRANCH_ORDER:
            venta[b] += float(reg[b]["venta"])
            tickets[b] += int(reg[b]["tickets"])
    return {b: round(venta[b], 2) for b in BRANCH_ORDER}, tickets


def acum_12m_por_sucursal(anio=None, mes=None, data=None):
    """Acumulado de los ultimos 12 meses cerrados hasta (anio, mes).
    Si no se pasa mes, usa el ultimo mes cargado (== ultimo cerrado). Lo usa el SEMANAL."""
    data = data if data is not None else cargar()
    if anio is None or mes is None:
        anio, mes = ultimo_mes(data)
    return _sumar(data, ventana_12m(anio, mes), f"acumulado 12 meses a {_clave(anio,mes)}")


def ytd_por_sucursal(anio, hasta_mes, data=None):
    """Acumulado del año calendario, enero..hasta_mes. Lo usa el CIERRE (actual y anterior)."""
    data = data if data is not None else cargar()
    claves = [(anio, m) for m in range(1, hasta_mes + 1)]
    return _sumar(data, claves, f"acumulado del año {anio} (ene..{hasta_mes:02d})")


# --- Control de conciliacion (dos caminos, un numero) ---------------------------
def conciliar_con_historial(anio, mes, historial):
    """Si el historial diario cubre el mes completo, su suma tiene que dar igual al
    registro mensual. historial = dict {'YYYY-MM-DD': {suc: {'venta','tickets'}|float}}.
    Devuelve texto para el log, o None si no habia con que conciliar. Aborta si no cierra."""
    from calendar import monthrange

    data = cargar()
    k = _clave(anio, mes)
    if k not in data:
        return None
    suma = {b: 0.0 for b in BRANCH_ORDER}
    for d in range(1, monthrange(anio, mes)[1] + 1):
        dia = date(anio, mes, d).isoformat()
        if dia not in historial:
            return (f"[CONCILIACION mensual] Omitida: el historial no cubre todo {k} "
                    f"(falta al menos {dia}).")
        for b in BRANCH_ORDER:
            v = historial[dia].get(b, 0.0)
            suma[b] += float(v["venta"]) if isinstance(v, dict) else float(v)
    difs = {b: round(suma[b] - data[k][b]["venta"], 2)
            for b in BRANCH_ORDER if abs(suma[b] - data[k][b]["venta"]) > 0.01}
    if difs:
        det = " | ".join(f"{b}: {v:+,.2f}" for b, v in difs.items())
        raise SystemExit(
            f"[ERROR] CONCILIACION MENSUAL FALLIDA en {k}: el master mensual y el "
            f"historial diario no cierran:\n        {det}\n"
            f"        Alguno tiene el mes mal cargado. NO se manda hasta entender por que."
        )
    return f"[CONCILIACION mensual] OK: {k} coincide con el historial al centavo."


# --- Backfill (una sola vez) ----------------------------------------------------
def backfill():
    """Reconstruye data/mensual_master.json desde:
      - data/Ventas_Master_2025.xlsx (diario, jun-dic 2025), agregado a mensual
      - todos los exports mensuales de TouchBistro que encuentre en ./backfill_exports/
        (formato multi-venue-sales-summary-YYYY-MM-01-YYYY-MM-DD.xlsx)
    Ajusta las rutas si tus exports estan en otro lado."""
    import glob
    from datetime import datetime
    import openpyxl
    from report import parse_excel_full, VENUE_MAP
    from generar_acum_ant import CODE_TO_TBKEY

    def _as_date(v):
        from datetime import datetime as dt
        if isinstance(v, dt):
            return v.date()
        if isinstance(v, date):
            return v
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()

    men = {}
    # (a) master diario 2025 -> mensual
    wb = openpyxl.load_workbook(BASE / "data" / "Ventas_Master_2025.xlsx", data_only=True)
    ws = wb["Por Dia y Local"]
    h = [str(c) for c in next(ws.iter_rows(values_only=True))]
    cl, cf, cs, ct = h.index("Local"), h.index("Fecha"), h.index("Sales"), h.index("Tickets")
    cod2b = {c: VENUE_MAP[k] for c, k in CODE_TO_TBKEY.items()}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[cl] is None:
            continue
        cod = str(r[cl]).strip()[:3]
        if cod not in cod2b:
            continue
        f = _as_date(r[cf]); key = _clave(f.year, f.month); b = cod2b[cod]
        d = men.setdefault(key, {bb: {"venta": 0.0, "tickets": 0} for bb in BRANCH_ORDER})
        d[b]["venta"] += float(str(r[cs]).replace(",", "").replace("$", "") or 0)
        d[b]["tickets"] += int(float(str(r[ct]).replace(",", "") or 0))
    # (b) exports mensuales
    for f in sorted(glob.glob(str(BASE / "backfill_exports" / "multi-venue-sales-summary-*.xlsx"))):
        ini, fin, v, t = parse_excel_full(f)
        men[_clave(ini.year, ini.month)] = {b: {"venta": round(v[b], 2), "tickets": t[b]} for b in BRANCH_ORDER}

    for k in men:
        men[k] = {b: {"venta": round(men[k][b]["venta"], 2), "tickets": int(men[k][b]["tickets"])}
                  for b in BRANCH_ORDER}
    guardar(men)
    print(f"[OK] mensual_master.json backfilleado con {len(men)} meses: "
          f"{sorted(men)[0]} -> {sorted(men)[-1]}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "backfill":
        backfill()
    else:
        data = cargar()
        ms = meses_disponibles(data)
        print(f"master mensual: {len(ms)} meses, {ms[0]} -> {ms[-1]}")
        v12, _ = acum_12m_por_sucursal(data=data)
        print(f"acum 12 meses (a {ultimo_mes(data)}): ${sum(v12.values()):,.2f}")
