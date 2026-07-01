"""
generar_acum2025.py
-------------------
Arma Acumulado_interanual.xlsx (comparativo del anio anterior) a partir del
master de ventas 2025, SIN tocar report.py.

Que hace:
  1. Lee el dia de corte del titulo de Ventas_ayer.xlsx (ej. 2026-06-30).
  2. Suma del master (data/Ventas_Master_2025.xlsx) la venta de cada local
     desde el 1ro del mismo mes del anio anterior hasta ese dia (2025-06-01..2025-06-30).
  3. Escribe Acumulado_interanual.xlsx con el MISMO formato que TouchBistro:
     fila 0 = titulo con rango, fila 1 = header, filas 2+ = local y Net Sales en col C.
     Usa los nombres estilo "#00X ..." para que report.py los reconozca (VENUE_MAP).

Falla RUIDOSA: si el master no tiene datos para ese mes, corta con error.
"""
import re
import sys
from pathlib import Path
from datetime import date, datetime

import openpyxl

BASE = Path(__file__).parent
VENTAS = BASE / "Ventas_ayer.xlsx"
MASTER = BASE / "data" / "Ventas_Master_2025.xlsx"
SALIDA = BASE / "Acumulado_interanual.xlsx"

# El master usa "00X - Nombre"; lo mapeo por el numero de local a la sucursal,
# y a la sucursal le asigno el nombre estilo TouchBistro que report.py espera.
CODE_TO_TBKEY = {
    "001": "#001 Florida Mall Orlando FL",
    "002": "#002 American Dream Mall NJ",
    "003": "#003 Sawgrass Mills Mall FL",
    "004": "#004 Weston Town Center FL",
    "005": "#005 Vineland Orlando FL",
    "006": "#006 Aventura, FL",
}
# Orden de salida (mismo criterio de codigo)
ORDEN = ["001", "002", "003", "004", "005", "006"]


def _to_float(v):
    if v is None:
        return 0.0
    return float(str(v).replace(",", "").replace("$", "").strip() or 0)


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def dia_de_corte():
    """Lee la fecha fin del titulo de Ventas_ayer.xlsx."""
    wb = openpyxl.load_workbook(VENTAS, data_only=True)
    ws = wb.active
    title = str(next(ws.iter_rows(values_only=True))[0] or "")
    m = re.search(r"(\d{4}-\d{2}-\d{2})\s*/\s*(\d{4}-\d{2}-\d{2})", title)
    if not m:
        raise ValueError(f"No pude leer la fecha de Ventas_ayer.xlsx: {title!r}")
    return datetime.strptime(m.group(2), "%Y-%m-%d").date()


def acumular_2025(corte):
    """Suma del master la columna 'Sales' por local, desde el 1ro del mes
    del anio anterior hasta el mismo dia de corte."""
    ini = date(corte.year - 1, corte.month, 1)
    fin = date(corte.year - 1, corte.month, corte.day)

    wb = openpyxl.load_workbook(MASTER, data_only=True)
    ws = wb["Por Dia y Local"]
    header = [str(c) for c in next(ws.iter_rows(values_only=True))]
    col_local = header.index("Local")
    col_fecha = header.index("Fecha")
    col_sales = header.index("Sales")

    acum = {cod: 0.0 for cod in ORDEN}
    hubo_datos = False
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[col_local] is None:
            continue
        cod = str(r[col_local]).strip()[:3]
        if cod not in acum:
            continue
        f = _as_date(r[col_fecha])
        if ini <= f <= fin:
            acum[cod] += _to_float(r[col_sales])
            hubo_datos = True

    if not hubo_datos:
        raise SystemExit(
            f"[ERROR] El master 2025 no tiene datos para {ini}..{fin}. "
            f"Cargale ese mes al master antes de correr."
        )
    return ini, fin, acum


def escribir(ini, fin, acum):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Summary"
    titulo = f"multi-venue - Sales Summary - {ini.isoformat()}/{fin.isoformat()}"
    ws.append([titulo, None, None, None, None, None])
    ws.append(["Venue Name", "Gross Sales", "Net Sales", "Discounts", "Voids", "Bill Count"])
    total = 0.0
    for cod in ORDEN:
        v = round(acum[cod], 2)
        total += v
        # Net Sales va en la columna C (index 2); Gross lo repito, no se usa.
        ws.append([CODE_TO_TBKEY[cod], v, v, 0, 0, 0])
    ws.append([f"REPORT SUMMARY ({len(ORDEN)} entries)", total, total, 0, 0, 0])
    wb.save(SALIDA)


if __name__ == "__main__":
    corte = dia_de_corte()
    ini, fin, acum = acumular_2025(corte)
    escribir(ini, fin, acum)
    tot = sum(acum.values())
    print(f"OK - Acumulado 2025 {ini}..{fin} generado. Total: ${tot:,.2f}")
